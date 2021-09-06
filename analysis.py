#!/usr/bin/env python
# -*- coding: utf-8 -*-
#

"""
Filename: analysis.py
Date: May 9, 2021 - 09:42
Name: Anonymous
Description:
    -  Analyze cic-ids2017 and lycos datasets: features importance,
    traditional ML algorithms

"""
import logging
import pandas as pd
import numpy as np
import os
import deep_learning as dl
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_selection import RFE
from sklearn.metrics import confusion_matrix
from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis, LinearDiscriminantAnalysis
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import LinearSVC
import tensorflow as tf

LOG_PATH = "./python_logs/"
ISCX_DATASET_PATH = "./datasets/cic-ids2017/"
LYCOS_DATASET_PATH = "./datasets/lycos-ids2017/"


FEAT_LIST_ISCX = {
    'flow_id': 0,
    'src_addr': 1,
    'src_port': 2,
    'dst_addr': 3,
    'dst_port': 4,
    'prot': 5,
    'timestamp': 6,
    'flow_duration': 7,
    'tot_fwd_pkt': 8,
    'tot_bwd_pkt': 9,
    'fwd_pkt_len_tot': 10,
    'bwd_pkt_len_tot': 11,
    'fwd_pkt_len_max': 12,
    'fwd_pkt_len_min': 13,
    'fwd_pkt_len_mean': 14,
    'fwd_pkt_len_std': 15,
    'bwd_pkt_len_max': 16,
    'bwd_pkt_len_min': 17,
    'bwd_pkt_len_mean': 18,
    'bwd_pkt_len_std': 19,
    'flow_byte_per_s': 20,
    'flow_pkt_per_s': 21,
    'flow_iat_mean': 22,
    'flow_iat_std': 23,
    'flow_iat_max': 24,
    'flow_iat_min': 25,
    'fwd_iat_tot': 26,
    'fwd_iat_mean': 27,
    'fwd_iat_std': 28,
    'fwd_iat_max': 29,
    'fwd_iat_min': 30,
    'bwd_iat_tot': 31,
    'bwd_iat_mean': 32,
    'bwd_iat_std': 33,
    'bwd_iat_max': 34,
    'bwd_iat_min': 35,
    'fwd_psh_flag': 36,
    'bwd_psh_flag': 37,
    'fwd_urg_flag': 38,
    'bwd_urg_flag': 39,
    'fwd_head_len': 40,
    'bwd_head_len': 41,
    'fwd_pkt_per_s': 42,
    'bwd_pkt_per_s': 43,
    'pkt_len_min': 44,
    'pkt_len_mean': 45,
    'pkt_len_std': 46,
    'pkt_len_max': 47,
    'pkt_len_var': 48,
    'fin_flag_cnt': 49,
    'syn_flag_cnt': 50,
    'rst_flag_cnt': 51,
    'psh_flag_cnt': 52,
    'ack_flag_cnt': 53,
    'urg_flag_cnt': 54,
    'cwe_flag_cnt': 55,
    'ece_flag_cnt': 56,
    'down_up_ratio': 57,
    'avg_pkt_size': 58,
    'avg_fwd_segm_size': 59,
    'avg_bwd_segm_size': 60,
    'fwd_head_len_bis': 61,
    'avg_fwd_byte_per_bulk': 62,
    'avg_fwd_pkt_per_bulk': 63,
    'avg_fwd_bulk_rate': 64,
    'avg_bwd_byte_per_bulk': 65,
    'avg_bwd_pkt_per_bulk': 66,
    'avg_bwd_bulk_rate': 67,
    'subflow_fwd_pkt': 68,
    'subflow_fwd_byte': 69,
    'subflow_bwd_pkt': 70,
    'subflow_bwd_byte': 71,
    'init_win_bytes_fwd': 72,
    'init_win_bytes_bwd': 73,
    'act_data_pkt_fwd': 74,
    'min_seg_size_forward': 75,
    'active_mean': 76,
    'active_std': 77,
    'active_max': 78,
    'active_min': 79,
    'idle_mean': 80,
    'idle_std': 81,
    'idle_max': 82,
    'idle_min': 83,
    'label': 84
}

FEAT_LIST_LYCOS = {
    'flow_id': 0,
    'src_addr': 1,
    'src_port': 2,
    'dst_addr': 3,
    'dst_port': 4,
    'ip_prot': 5,
    'timestamp': 6,
    'flow_duration': 7,
    'down_up_ratio': 8,
    'pkt_len_max': 9,
    'pkt_len_min': 10,
    'pkt_len_mean': 11,
    'pkt_len_var': 12,
    'pkt_len_std': 13,
    'bytes_per_s': 14,
    'pkt_per_s': 15,
    'fwd_pkt_per_s': 16,
    'bwd_pkt_per_s': 17,
    'fwd_pkt_cnt': 18,
    'fwd_pkt_len_tot': 19,
    'fwd_pkt_len_max': 20,
    'fwd_pkt_len_min': 21,
    'fwd_pkt_len_mean': 22,
    'fwd_pkt_len_std': 23,
    'fwd_pkt_hdr_len_tot': 24,
    'fwd_pkt_hdr_len_min': 25,
    'fwd_non_empty_pkt_cnt': 26,
    'bwd_pkt_cnt': 27,
    'bwd_pkt_len_tot': 28,
    'bwd_pkt_len_max': 29,
    'bwd_pkt_len_min': 30,
    'bwd_pkt_len_mean': 31,
    'bwd_pkt_len_std': 32,
    'bwd_pkt_hdr_len_tot': 33,
    'bwd_pkt_hdr_len_min': 34,
    'bwd_non_empty_pkt_cnt': 35,
    'iat_max': 36,
    'iat_min': 37,
    'iat_mean': 38,
    'iat_std': 39,
    'fwd_iat_tot': 40,
    'fwd_iat_max': 41,
    'fwd_iat_min': 42,
    'fwd_iat_mean': 43,
    'fwd_iat_std': 44,
    'bwd_iat_tot': 45,
    'bwd_iat_max': 46,
    'bwd_iat_min': 47,
    'bwd_iat_mean': 48,
    'bwd_iat_std': 49,
    'active_max': 50,
    'active_min': 51,
    'active_mean': 52,
    'active_std': 53,
    'idle_max': 54,
    'idle_min': 55,
    'idle_mean': 56,
    'idle_std': 57,
    'flag_syn': 58,
    'flag_fin': 59,
    'flag_rst': 60,
    'flag_ack': 61,
    'flag_psh': 62,
    'fwd_flag_psh': 63,
    'bwd_flag_psh': 64,
    'flag_urg': 65,
    'fwd_flag_urg': 66,
    'bwd_flag_urg': 67,
    'flag_cwr': 68,
    'flag_ece': 69,
    'fwd_bulk_bytes_mean': 70,
    'fwd_bulk_pkt_mean': 71,
    'fwd_bulk_rate_mean': 72,
    'bwd_bulk_bytes_mean': 73,
    'bwd_bulk_pkt_mean': 74,
    'bwd_bulk_rate_mean': 75,
    'fwd_subflow_bytes_mean': 76,
    'fwd_subflow_pkt_mean': 77,
    'bwd_subflow_bytes_mean': 78,
    'bwd_subflow_pkt_mean': 79,
    'fwd_TCP_init_win_bytes': 80,
    'bwd_TCP_init_win_bytes': 81,
    'label': 82,
}

LYCOS_LABEL_DICT = {
    "benign": 0,
    "bot": 1,
    "ddos": 2,
    "dos_goldeneye": 3,
    "dos_hulk": 4,
    "dos_slowhttptest": 5,
    "dos_slowloris": 6,
    "ftp_patator": 7,
    "heartbleed": 8,
    "portscan": 9,
    "ssh_patator": 10,
    "webattack_bruteforce": 11,
    "webattack_sql_injection": 12,
    "webattack_xss": 13,
}


def initialize_logger(output_dir):
    """
    Initialize logger so that errors are logged in error.log, all messages
    are logged in all.log printed in console

    Parameters
    ----------
    output_dir: str
        String corresponding to the path of the log file

    Returns
    -------
    None
    """
    pyfile = os.path.splitext(os.path.split(__file__)[1])[0]
    filename = pyfile + "_info.log"
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    # create console handler and set level to info
    handler = logging.StreamHandler()
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter("[%(levelname)s] %(asctime)s - "
                                  "%(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    # create debug file handler and set level to debug
    handler = logging.FileHandler(os.path.join(output_dir, filename), "w")
    handler.setLevel(logging.DEBUG)
    formatter = logging.Formatter("[%(levelname)s] %(asctime)s - "
                                  "%(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    handler.setFormatter(formatter)
    logger.addHandler(handler)


def read_csv(filename):
    """
    load csv file into a DataFrame

    Parameters
    ----------
    filename: str
        file name of the CSV that is loaded

    Returns
    -------
    df: DataFrame
        Pandas DataFrame containing data from CSV
    """
    logging.info("Loading CSV {}".format(filename))
    df = pd.read_csv(filename, encoding="ISO-8859-1", low_memory=False,
                     index_col=False)
    n_rows, n_cols = df.shape
    logging.info("Loaded CSV {} - shape: [{},{}]".format(filename,
                                                         n_rows, n_cols))
    return df


def write_xls(df, filename):
    """
    Write
    Parameters
    ----------
    df: DataFrame
        Contains data to stoe in Excel file
    filename: str
        name of the Excel file

    Returns
    -------
    None
    """
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=None, header=True):
        ws.append(r)
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    wb.save(filename)


def load_dataset(filename, file_format, dataset_name):
    """
    Load extract of CIC2017 dataset from a pre-processed CSV file
    Included processing: normalization, randomization, split

    Parameter
    ---------
    filename: string
        name of the CSV file to be loaded
    file_format: string
        supported format: 'csv' and 'parquet'
    dataset_name: string
        valid strings are 'cic-ids2017' and 'lycos-ids2017'

    Return
    ------
    dataset: Dataset
        Returns a Dataset object
    """
    # Use size_percent to reduce dataset
    size_percent = 100
    # load dataset in a pandas DataFrame
    logging.info("Loading dataset from {}".format(filename))
    ds = dl.Dataset()
    ds.from_file(filename, file_format, size_percent)
    logging.info("Dataset shape: {}".format(ds.data.shape))

    # Select labels from DataFrame
    ds.select_dfcol_as_label('label', False)
    if dataset_name == 'cic-ids2017':
        drop_list = [
            'flow_byte_per_s', 'flow_pkt_per_s',
            'bwd_psh_flag', 'bwd_urg_flag',
            'avg_fwd_byte_per_bulk', 'avg_fwd_pkt_per_bulk',
            'avg_fwd_bulk_rate', 'avg_bwd_byte_per_bulk',
            'avg_bwd_pkt_per_bulk', 'avg_bwd_bulk_rate',
            'flow_id', 'timestamp',
            'src_addr', 'src_port', 'dst_addr',
            'label'
        ]
        ds.drop_dfcol(drop_list)
    else:
        drop_list = [
            'flag_urg', 'fwd_flag_urg', 'bwd_flag_urg',
            'flow_id', 'timestamp',
            'src_addr', 'src_port', 'dst_addr',
            'label'
        ]
        ds.drop_dfcol(drop_list)
    logging.info("Dataset shape after feature selection: {}".
                 format(ds.data.shape))
    return ds


def load_files(dataset_name):
    """
    Load training, cross-val and test sets

    Parameter
    ---------
    dataset_name: string
        valid strings are 'cic-ids2017' and 'lycos-ids2017'

    Returns
    -------
    ds_train: Dataset
        Dataset for training
    ds_cv: Dataset
        Dataset for cross-validation
    ds_test: Dataset
        Dataset for test
    """
    if dataset_name == "cic-ids2017":
        path = ISCX_DATASET_PATH
    elif dataset_name == "lycos-ids2017":
        path = LYCOS_DATASET_PATH
    else:
        logging.debug("Wrong dataset name")
        raise Exception("[ERROR] Wrong dataset name")

    filename_train = path + 'train_set.parquet'
    filename_cv = path + 'crossval_set.parquet'
    filename_test = path + 'test_set.parquet'
    logging.info("Loading data sets ...")
    # Load dataset in a pandas DataFrame
    ds_train = load_dataset(filename_train, 'parquet', dataset_name)
    ds_cv = load_dataset(filename_cv, 'parquet', dataset_name)
    ds_test = load_dataset(filename_test, 'parquet', dataset_name)
    return ds_train, ds_cv, ds_test


def normalize(ds_train, ds_cv, ds_test, data_type, clipping_val):
    """
    Normalization of datasets including clipping

    Parameters
    ----------
    ds_train: Dataset
        Training set
    ds_cv: Dataset
        Cross-validation set
    ds_test: Dataset
        Test set
    data_type: str
        Type of data. Accepted strings are 'float64', 'float32', 'int8'
    clipping_val: int
        value used for clipping

    Returns
    -------
    norm_train: Dataset
        Normalized training set
    norm_cv: Dataset
        Normalized cross-validation set
    norm_test: Dataset
        Normalized test set
    """
    logging.info("Normalizing training set ...")
    normalizer = dl.Normalization()
    ds_train.data = normalizer.fit_and_transform(ds_train.data,
                                                 method='z_score_std',
                                                 per_col_scaler=True)
    if clipping_val != 0:
        ds_train.data = normalizer.clip(ds_train.data,
                                        clip_val_low=-clipping_val,
                                        clip_val_high=clipping_val)
    logging.info("Normalizing crossval set ...")
    ds_cv.data = normalizer.transform(ds_cv.data)
    if clipping_val != 0:
        ds_cv.data = normalizer.clip(ds_cv.data,
                                     clip_val_low=-clipping_val,
                                     clip_val_high=clipping_val)
    logging.info("Normalizing test set ...")
    ds_test.data = normalizer.transform(ds_test.data)
    if clipping_val != 0:
        ds_test.data = normalizer.clip(ds_test.data,
                                       clip_val_low=-clipping_val,
                                       clip_val_high=clipping_val)
    if data_type == 'float64':
        ds_train.data = ds_train.data.astype(dtype='float64')
        ds_cv.data = ds_cv.data.astype(dtype='float64')
        ds_test.data = ds_test.data.astype(dtype='float64')
    elif data_type == 'float32':
        ds_train.data = ds_train.data.astype(dtype='float32')
        ds_cv.data = ds_cv.data.astype(dtype='float32')
        ds_test.data = ds_test.data.astype(dtype='float32')
    else:
        logging.debug("Normalization data_type not supported")
        raise Exception("[ERROR] normalization data_type not supported")
    return ds_train, ds_cv, ds_test


def print_metrics(ds, cost, cm):
    """
    Calculate and print some metrics on predictions

    Parameter
    ---------
    ds: Dataset
        Dataset
    cost: float
        overall cost
    cm: ndarray
        confusion matrix (row=true labels ; col=predicted labels)

    Return
    ------
    tp: int
        number of true positive
    fp: int
        number of false positive
    fn: int
        number of false negative
    tn: int
        number of true negative
    tnr: float
        true negative rate
    fpr: float
        false positive rate
    recall: float
        recall rate
    precision: float
        precision rate
    accuracy: float
        accuracy rate
    f1_score: float
        Harmonic mean of precision and recall
    MCC: float
        Matthews correlation coefficient
    """
    # create a simplified confusion matrix for multi-class classification
    simpl_cm = np.zeros(3*len(ds.label_dict))
    simpl_cm = simpl_cm.reshape((len(ds.label_dict), 3))
    simpl_cm[0][0] = cm[0][0]            # benign classified as benign
    simpl_cm[0][1] = 0                   # not applicable
    simpl_cm[0][2] = np.sum(cm[:1, 1:])  # benign classified as attack
    for i in range(1, cm.shape[0]):
        # Attack i classified as benign
        simpl_cm[i][0] = cm[i][0]
        # attack i correctly classified
        simpl_cm[i][1] = cm[i][i]
        # attack i classified as another attack
        simpl_cm[i][2] = np.sum(cm[i][1:i]) + np.sum(cm[i][i + 1:])
    simpl_cm = np.int_(simpl_cm)

    # create a confusion matrix for a binary classification (Normal vs Attack)
    binary_cm = np.zeros(4)
    binary_cm = binary_cm.reshape((2, 2))
    # True positive: attacks correctly detected
    # tp = np.sum(np.diagonal(cm[1:, 1:]))
    tp = np.sum(simpl_cm[1:, 1:])
    # True negative: normal traffic correctly detected
    tn = cm[0][0]
    # False positive: normal traffic detected as attacks
    fp = np.sum(simpl_cm[0, 1:])
    # False negative: attacks detected as normal traffic
    fn = np.sum(simpl_cm[1:, 0])
    binary_cm[0][0] = tp
    binary_cm[0][1] = fn
    binary_cm[1][0] = fp
    binary_cm[1][1] = tn
    binary_cm = np.int_(binary_cm)

    # calculate metrics
    accuracy = (tp+tn) / (tp+tn+fp+fn)
    precision = tp / (tp+fp)
    # recall = sensitivity
    recall = tp / (tp+fn)
    # f1-score is the harmonic mean of precision and recall.
    f1_score = (2*precision*recall) / (precision + recall)
    # True negative rate = specificity
    tnr = tn / (tn+fp)
    # False positive rate = fp / (tn+fp)
    fpr = 1 - tnr
    # Matthews correlation coefficient
    # MCC = (tp*tn-fp*fn)/math.sqrt((tp+fp)*(tp+fn)*(tn+fp)*(tn+fn))
    mcc = 0
    mcc_exists = False
    if tp+fp != 0 and tp+fn != 0 and tn+fp != 0 and tn+fn != 0:
        mcc_exists = True
        mcc = np.float64(tp)*np.float64(tn)-np.float64(fp)*np.float64(fn)
        mcc = mcc / np.float64(np.sqrt(tp+fp))
        mcc = mcc / np.float64(math.sqrt(tp+fn))
        mcc = mcc / np.float64(math.sqrt(tn+fp))
        mcc = mcc / np.float64(math.sqrt(tn+fn))

    logging.info(" ********** {}  **********".format(ds.data_file))
    logging.info("Simplified confusion matrix")
    logging.info("row=true labels (alphabet. order)")
    logging.info("col=Benign | Correct Attack | Other Attack:")
    logging.info(simpl_cm)
    logging.info("Binary confusion matrix (row=true labels ; col=predicted labels)")
    logging.info("  Attack  Benign")
    logging.info(binary_cm)
    logging.info("overall cost: {}".format(cost))
    logging.info("true positive: {}".format(tp))
    logging.info("false positive: {}".format(fp))
    logging.info("false negative: {}".format(fn))
    logging.info("true negative: {}".format(tn))
    logging.info("true negative rate: {0:.6f}".format(tnr))
    logging.info("false positive rate: {0:.6f}".format(fpr))
    logging.info("recall: {0:.6f}".format(recall))
    logging.info("precision: {0:.6f}".format(precision))
    logging.info("accuracy: {0:.6f}".format(accuracy))
    logging.info("f1_score: {0:.6f}".format(f1_score))
    if mcc_exists is True:
        logging.info("MCC: {0:.6f}".format(mcc))

    return (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
            mcc)


def get_performances(nn, ds_train, ds_cv, ds_test):
    """
    Create and train a simple MLP

    Parameters
    ----------
    nn: NeuralNetwork
        Neural network to train
    ds_train: pandas DataFrame
        Training set
    ds_cv: pandas DataFrame
        Cross-validation set
    ds_test: pandas DataFrame
        Test set
    Returns
    -------
    None
    """
    # get metrics for training set
    logging.info("Calculating training metrics ...")
    (train_loss, train_acc, train_pred, train_cm) = nn.get_metrics(ds_train)
    logging.info("Training cost: {}".format(train_loss))
    logging.info("Training accuracy: {}".format(train_acc))
    logging.info("Training confusion matrix:")
    logging.info(train_cm)
    # get metrics for crossval set
    logging.info("Calculating crossval metrics ...")
    (cv_loss, cv_acc, cv_pred, cv_cm) = nn.get_metrics(ds_cv)
    logging.info("Crossval cost: {}".format(cv_loss))
    logging.info("Crossval accuracy: {}".format(cv_acc))
    logging.info("Crossval confusion matrix:")
    logging.info(cv_cm)
    # get metrics for test set
    logging.info("Calculating test metrics ...")
    t1 = datetime.now()
    (test_loss, test_acc, test_pred, test_cm) = nn.get_metrics(ds_test)
    t2 = datetime.now()
    logging.info("Test cost: {}".format(test_loss))
    logging.info("Test accuracy: {}".format(test_acc))
    logging.info("Test confusion matrix:")
    logging.info(test_cm)
    print_metrics(ds_train, train_loss, train_cm)
    print_metrics(ds_cv, cv_loss, cv_cm)
    print_metrics(ds_test, test_loss, test_cm)
    t_train = t2 - t1
    logging.info("MLP prediction time: {}".format(t_train))


def create_nn(n_inputs, n_outputs, name,
              activation='selu',
              dropout_rate=0.5, l2_reg=0.0,
              silent_console=False):
    """
    Declare and initialize neural network. Initialize default values for
    hyper-parameters.

    Parameters
    ----------
    n_inputs: int
        number of inputs (features) for the neural network
    n_outputs: int
        number of outputs (classes) for the neural network
    name: str
        string containing neural network name
    activation: str
        string indicating activation function to use
        supported functions are 'selu', 'sigmoid', 'relu'
    dropout_rate: float
        drop-out rate used for regularization
        default value 0 (no regularization)
    l2_reg: float
        L2 norm factor used for regularization
        default value 0 (no regularization)
    silent_console: bool
        When True, no message is put in the console

    Returns
    -------
    nn: NeuralNetwork
        Neural network that has been created and initialized
    """
    if silent_console is False:
        logging.info("Creating neural network ...")
    if activation == 'selu':
        act_func = 'selu'
    elif activation == 'relu':
        act_func = 'relu'
    elif activation == 'sigmoid':
        act_func = 'sigmoid'
    else:
        logging.debug("[ERROR] activation function not supported")
        raise Exception("[ERROR] activation function not supported")
    nn_desc = dl.NeuralNetworkDescriptor(name)
    layer1 = dl.Layer(name='layer1_{}'.format(act_func),
                      layer_type='dense',
                      n_inputs=n_inputs,
                      n_nodes=256,
                      activation_name=act_func)
    nn_desc.topology.add_layer(layer1)
    l1_drop = dl.Layer(name='dropout1',
                       layer_type='dropout',
                       dropout_rate=dropout_rate,
                       n_nodes=256,
                       activation_name=None)
    nn_desc.topology.add_layer(l1_drop)
    layer2 = dl.Layer(name='layer2_{}'.format(act_func),
                      layer_type='dense',
                      n_nodes=256,
                      activation_name=act_func)
    nn_desc.topology.add_layer(layer2)
    l2_drop = dl.Layer(name='dropout2',
                       layer_type='dropout',
                       n_nodes=256,
                       dropout_rate=dropout_rate,
                       activation_name=None)
    nn_desc.topology.add_layer(l2_drop)
    layer3 = dl.Layer(name='layer3_softmax',
                      layer_type='dense',
                      n_nodes=n_outputs,
                      activation_name='softmax')
    # prepare output bias according to probability of each class, assuming a
    # softmax activation at output layer
    output_bias = [0.500756854747337,
                   0.0008316338092,
                   0.108409245411285,
                   0.007663720824836,
                   0.180135961930659,
                   0.005513256288239,
                   0.006428733288013,
                   0.004534330387492,
                   0.000011330160883,
                   0.180070246997507,
                   0.003351461590755,
                   0.001540901880807,
                   0.000013596193066,
                   0.000738726489916]
    layer3.output_bias = np.log(output_bias)
    nn_desc.topology.add_layer(layer3)
    nn_desc.hyper_param.optimizer_name = 'Adam'
    # params: learning_rate, Beta1, Beta2
    nn_desc.hyper_param.optim_params = [0.001, 0.9, 0.999]
    nn_desc.hyper_param.set_optimizer_params()
    nn_desc.hyper_param.epochs = 10
    nn_desc.hyper_param.l1_reg = 0.0
    nn_desc.hyper_param.l2_reg = 0.0
    # nn_desc.hyper_param.keep_prob = 0.5
    nn_desc.hyper_param.batch_size = 32
    nn_desc.hyper_param.eval_print_period = 1
    nn = dl.NeuralNetwork(nn_descriptor=nn_desc)
    return nn


def train_nn(nn, ds_train, ds_cv):
    """
    Create and train a simple MLP

    Parameters
    ----------
    nn: NeuralNetwork
        Neural network to train
    ds_train: pandas DataFrame
        Training set
    ds_cv: pandas DataFrame
        Cross-validation set

    Returns
    -------
    None
    """
    # training phase
    logging.info("Training neural network ...")
    t1 = datetime.now()
    nn.fit(ds_train, ds_cv)
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("MLP training time: {}".format(t_train))
    logging.info("Training done")


def measure_perf(ds_train, ds_cv, ds_test, dataset_name):
    """
    Measure performance of ML algorithms on normalized
    dataset:
    LDA, QDA, Decision Tree, Random Forest, SVM, k-NN, MLP

    Parameters
    ----------
    ds_train: pandas DataFrame
        Training set
    ds_cv: pandas DataFrame
        Cross-validation set
    ds_test: pandas DataFrame
        Test set
    dataset_name: str
        valid strings are 'cic-ids2017' and 'lycos-ids2017'

    Returns
    -------
    None
    """
    # create a dataframe with metrics
    data = {'Metrics': ['TP', 'FP', 'FN', 'TN', 'TNR', 'FPR', 'Recall',
                        'Precision', 'Accuracy', 'F1_score', 'MCC',
                        'Training time', 'Prediction time']}
    df = pd.DataFrame(data)
    df.set_index('Metrics')

    logging.info("------------------------------------------")
    logging.info("MLP")
    # training
    logging.info("Training ...")
    tf.random.set_seed(2021)
    # set parameters
    n_epochs = 25
    activation = 'selu'
    if dataset_name == "lycos-ids2017":
        dropout_rate = 0.11859054480832958
        optim_params = [0.0003175219400915149,
                        0.9158410351985389,
                        0.999]
    else:
        dropout_rate = 0.20886178816521117
        optim_params = [0.0003590920789688343,
                        0.8662527597085028,
                        0.999]

    # adjust number of classes
    ds_train.n_classes = len(ds_train.labels.unique())
    ds_cv.n_classes = ds_train.n_classes
    ds_test.n_classes = ds_train.n_classes
    # create neural network
    mlp = create_nn(n_inputs=ds_train.n_features,
                    n_outputs=ds_train.n_classes,
                    activation=activation,
                    dropout_rate=dropout_rate,
                    name='mlp_'+dataset_name)
    logging.info("Model name: {}".format(mlp.name))
    logging.info("n_epochs: {}".format(n_epochs))
    logging.info("activation function: {}".format(activation))
    logging.info("dropout_rate: {}".format(dropout_rate))
    logging.info("optim_params: {}".format(optim_params))
    mlp.model.summary()
    mlp.nn_desc.hyper_param.epochs = n_epochs
    mlp.nn_desc.hyper_param.optim_params = optim_params
    t1 = datetime.now()
    train_nn(mlp, ds_train, ds_cv)
    # retrieve best recorded epoch instead of the last epoch
    mlp.restore (ds_cv)
    # Predictions
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("Inference ...")
    (loss, acc, pred, cm) = mlp.get_metrics(ds_test)
    t3 = datetime.now()
    t_pred = t3 - t2
    logging.info("Results:")
    (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
     mcc) = print_metrics(ds_test, 0, cm)
    df['MLP'] = [tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy,
                 f1_score, mcc, t_train.total_seconds(), t_pred.total_seconds()]

    logging.info("------------------------------------------")
    logging.info("Linear Discriminant Analysis")
    lda = LinearDiscriminantAnalysis()
    # training
    t1 = datetime.now()
    logging.info("Training ...")
    lda.fit(ds_train.data, ds_train.labels)
    # Predictions
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("Inference ...")
    y_pred = lda.predict(ds_test.data)
    t3 = datetime.now()
    t_pred = t3 - t2
    logging.info("Results:")
    cm = confusion_matrix(ds_test.labels, y_pred)
    (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
     mcc) = print_metrics(ds_test, 0, cm)
    df['LDA'] = [tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy,
                 f1_score, mcc, t_train.total_seconds(), t_pred.total_seconds()]

    logging.info("------------------------------------------")
    logging.info("Quadratic Discriminant Analysis")
    qda = QuadraticDiscriminantAnalysis()
    # training
    t1 = datetime.now()
    logging.info("Training ...")
    qda.fit(ds_train.data, ds_train.labels)
    # Predictions
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("Inference ...")
    y_pred = qda.predict(ds_test.data)
    t3 = datetime.now()
    t_pred = t3 - t2
    logging.info("Results:")
    cm = confusion_matrix(ds_test.labels, y_pred)
    (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
     mcc) = print_metrics(ds_test, 0, cm)
    df['QDA'] = [tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy,
                 f1_score, mcc, t_train.total_seconds(), t_pred.total_seconds()]

    logging.info("------------------------------------------")
    logging.info("SVM")
    # declare classifier
    svm = LinearSVC(random_state=0, verbose=0)
    # training
    t1 = datetime.now()
    logging.info("Training ...")
    svm.fit(ds_train.data, ds_train.labels)
    # Predictions
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("Inference ...")
    y_pred = svm.predict(ds_test.data)
    t3 = datetime.now()
    t_pred = t3 - t2
    logging.info("Results:")
    cm = confusion_matrix(ds_test.labels, y_pred)
    (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
     mcc) = print_metrics(ds_test, 0, cm)
    df['SVM'] = [tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy,
                 f1_score, mcc, t_train.total_seconds(), t_pred.total_seconds()]

    logging.info("------------------------------------------")
    logging.info("k-Nearest Neighbors")
    knn = KNeighborsClassifier(n_neighbors=5, n_jobs=1)
    # training
    t1 = datetime.now()
    logging.info("Training ...")
    knn.fit(ds_train.data, ds_train.labels)
    # Predictions
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("Inference ...")
    y_pred = knn.predict(ds_test.data)
    t3 = datetime.now()
    t_pred = t3 - t2
    logging.info("Results:")
    cm = confusion_matrix(ds_test.labels, y_pred)
    (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
     mcc) = print_metrics(ds_test, 0, cm)
    df['k-NN'] = [tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy,
                  f1_score, mcc, t_train.total_seconds(), t_pred.total_seconds()]

    logging.info("------------------------------------------")
    logging.info("Decision Tree")
    # declare classifier
    dt = DecisionTreeClassifier(random_state=0)
    # training
    t1 = datetime.now()
    logging.info("Training ...")
    dt.fit(ds_train.data, ds_train.labels)
    # Predictions
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("Inference ...")
    y_pred = dt.predict(ds_test.data)
    t3 = datetime.now()
    t_pred = t3 - t2
    logging.info("Results:")
    cm = confusion_matrix(ds_test.labels, y_pred)
    # print(cm)
    # print(classification_report(ds_test.labels, y_pred))
    (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
     mcc) = print_metrics(ds_test, 0, cm)
    df['DT'] = [tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy,
                f1_score, mcc, t_train.total_seconds(), t_pred.total_seconds()]

    logging.info("------------------------------------------")
    logging.info("Random Forest")
    # declare classifier
    rf = RandomForestClassifier(n_jobs=1, random_state=0, verbose=1)
    # training
    t1 = datetime.now()
    logging.info("Training ...")
    rf.fit(ds_train.data, ds_train.labels)
    # Predictions
    t2 = datetime.now()
    t_train = t2 - t1
    logging.info("Inference ...")
    y_pred = rf.predict(ds_test.data)
    t3 = datetime.now()
    t_pred = t3 - t2
    logging.info("Results:")
    cm = confusion_matrix(ds_test.labels, y_pred)
    # print(cm)
    # print(classification_report(ds_test.labels, y_pred))
    (tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy, f1_score,
     mcc) = print_metrics(ds_test, 0, cm)
    df['RF'] = [tp, fp, fn, tn, tnr, fpr, recall, precision, accuracy,
                f1_score, mcc, t_train.total_seconds(), t_pred.total_seconds()]

    # save results in excel file
    filename = './' + dataset_name + '_metrics.xls'
    write_xls(df, filename)
    logging.info(df)


def main():
    """
    Main program

    """
    # initialization
    initialize_logger(LOG_PATH)
    desired_width = 320
    pd.set_option('display.width', desired_width)
    np.set_printoptions(linewidth=desired_width)
    # np.random.seed(2021)
    np.random.seed(2021)

    # CIC-IDS2017 original dataset
    # Load dataset in a pandas DataFrame
    ds_train_cic, ds_cv_cic, ds_test_cic = load_files("cic-ids2017")
    # Normalize data
    ds_train_cic, ds_cv_cic, ds_test_cic = normalize(ds_train_cic, ds_cv_cic,
                                                     ds_test_cic,
                                                     data_type='float64',
                                                     clipping_val=0)
    # LYCOS-IDS2017 dataset
    # Load dataset in a pandas DataFrame
    ds_train_lycos, ds_cv_lycos, ds_test_lycos = load_files("lycos-ids2017")
    # Normalize data
    ds_train_lycos, ds_cv_lycos, ds_test_lycos = normalize(ds_train_lycos,
                                                           ds_cv_lycos,
                                                           ds_test_lycos,
                                                           data_type='float64',
                                                           clipping_val=0)

    measure_perf(ds_train_lycos, ds_cv_lycos, ds_test_lycos, "lycos-ids2017")
    measure_perf(ds_train_cic, ds_cv_cic, ds_test_cic, "cic-ids2017")

    logging.info("----Program finished----")


if __name__ == "__main__":
    main()
